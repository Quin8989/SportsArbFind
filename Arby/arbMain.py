import requests
import heapq
from collections import defaultdict
import copy
from openpyxl import load_workbook
global betsFound
betsFound = 0

# An api key is emailed to you when you sign up to a plan
# Get a free API key at https://api.the-odds-api.com/
API_KEY = ''

SPORT = 'upcoming' # use the sport_key from the /sports endpoint below, or use 'upcoming' to see the next 8 games across all sports

REGIONS = 'us,uk,eu' # uk | us | eu | au. Multiple can be specified if comma delimited

MARKETS = 'h2h' # h2h | spreads | totals. Multiple can be specified if comma delimited

ODDS_FORMAT = 'decimal' # decimal | american

DATE_FORMAT = 'iso' # iso | unix

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
#
# Get a list of in-season sports
# The sport 'key' from the response can be used to get odds in the next request
#
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
def getSports():
    sports_response = requests.get(
    'https://api.the-odds-api.com/v4/sports', 
    params={
        'api_key': API_KEY
    }
)


    if sports_response.status_code != 200:
        print(f'Failed to get sports: status_code {sports_response.status_code}, response body {sports_response.text}')

    else:
        jresp = sports_response.json()
        for ind in jresp:
            print(ind)
    
    
    # Check the usage quota
    print('Remaining requests', sports_response.headers['x-requests-remaining'])
    print('Used requests', sports_response.headers['x-requests-used'])



# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
#
# Get a list of live & upcoming games for the sport you want, along with odds for different bookmakers
# Deducts from the usage quota
# Usage quota cost = [number of markets specified] x [number of regions specified]
#
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

def getOdds():
    odds_response = requests.get(
    f'https://api.the-odds-api.com/v4/sports/{SPORT}/odds',
    params={
        'api_key': API_KEY,
        'regions': REGIONS,
        'markets': MARKETS,
        'oddsFormat': ODDS_FORMAT,
        'dateFormat': DATE_FORMAT,
    }
    )

    if odds_response.status_code != 200:
        print(f'Failed to get odds: status_code {odds_response.status_code}, response body {odds_response.text}')
        return

    else:
        arbList = []
        jresp = odds_response.json()
        for game in jresp:
            sport_key = game['sport_key']
            sport_title = game['sport_title']
            commence_time = game['commence_time']
            home_team = game['home_team']
            away_team = game['away_team']
            oddsMap = defaultdict(list)

            for bookmaker in game['bookmakers']:
                title = bookmaker['title']
                if title == 'Betfair' or title == 'GTBets' or title == 'Matchbook': # these books don't allow Canadian accounts
                    continue   
                for market in bookmaker['markets']:
                    for outcomes in market['outcomes']:
                        name = outcomes['name']
                        price = outcomes['price']
                        oddsMap[name].append(OddsInfo(price, name, title, sport_key, commence_time))   
                   
            for keys in oddsMap.values():
                keys.sort(key = lambda x : x.odds, reverse = True)

            arbList.append(Game(copy.deepcopy(sport_key), copy.deepcopy(commence_time), copy.deepcopy(oddsMap)))
            #print(" ")
            #print(oddsMap)
            oddsMap.clear()

        findArbs(arbList)
        print("Bets Found: {}".format(betsFound))
        #print(arbList)


        

    # Check the usage quota
    print('Remaining requests', odds_response.headers['x-requests-remaining'])
    print('Used requests', odds_response.headers['x-requests-used'])
    #print(jresp)

def findArbs(arbList):
    for game in arbList:
        #print(len(arbList))
        cond = True
        while cond == True: # while the sorted list still has profitable combos
            arbOpp = []
            for side in game.oddsMap: # for every possible betting side in game
                if len(game.oddsMap[side]) == 0:
                    cond = False
                    continue
                bet = game.oddsMap[side].pop(0) # remove the most attractive bet in side
                arbOpp.append(bet)
            arbPercent = calcArb(arbOpp)
            if arbPercent > 1:
                cond = False
            else:
                if arbPercent == 0:
                    continue
                global betsFound
                betsFound = betsFound + 1
                print(" ")
                print("Arb Ratio: {}".format(arbPercent))
                print("Sport: {}".format(arbOpp[0].sport))
                #print(arbOpp)
                for bet in arbOpp:
                    betRatio = (1/bet.odds)/(arbPercent)
                    print("Bet: {}, Odds: {}, Book: {}, HedgeRatio: {}".format(bet.side, bet.odds, bet.book, betRatio))

                    for i in range(2,34):
                        b_col = sheet.cell(row = i , column = 1)
                        if b_col.value == bet.book:
                            sheet.cell(row=i,column=2).value = sheet.cell(row=i,column=2).value + 1


                            

def calcArb(H):
    arbRatio = 0
    for x in H:
        arbRatio += 1/(x.odds)
    return arbRatio



class OddsInfo:
    def __init__(self, odds, side, book, sport, commence_time):
        self.odds = odds
        self.side = side
        self.book = book
        self.sport = sport
        self.commence_time = commence_time

    def __repr__(self):
        return "|| {} | {} | {} ||".format(self.odds, self.side, self.book)

class Game:
    def __init__(self, sport_title, commence_time, oddsMap):
        self.sport_title = sport_title
        self.commence_time = commence_time
        self.oddsMap = oddsMap
    
    def __repr__(self):
        return repr(self.oddsMap)
    

book = load_workbook('booksinfo.xlsx')
sheet = book['Sheet1']
#getSports()
getOdds()
book.save("booksinfo.xlsx")