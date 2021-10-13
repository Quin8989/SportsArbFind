from openpyxl import load_workbook
book = load_workbook('modTest.xlsx')

sheet = book['Sheet1']
for i in range(2,7):
    b_col = sheet.cell(row = i , column = 1)
    if b_col.value == 'shake':
        sheet.cell(row=i,column=2).value = sheet.cell(row=i,column=2).value + 1

book.save("modTest.xlsx")


